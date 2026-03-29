from dotenv import load_dotenv
load_dotenv()

import os
import secrets
import base64
import httpx
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from bson import ObjectId
from fastapi import FastAPI, HTTPException, Request, Response, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, Field
from motor.motor_asyncio import AsyncIOMotorClient
from contextlib import asynccontextmanager
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl

# ============ CONFIG ============
JWT_ALGORITHM = "HS256"
POKEMON_TCG_API = "https://api.pokemontcg.io/v2"

# ============ DATABASE ============
client = None
db = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global client, db
    mongo_url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
    db_name = os.environ.get("DB_NAME", "pokemon_classeur")
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]
    
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)
    await db.login_attempts.create_index("identifier")
    await db.cards.create_index([("user_id", 1), ("pokemon_name", 1)])
    
    # Seed admin
    await seed_admin()
    
    yield
    client.close()

app = FastAPI(title="Pokemon Classeur API", lifespan=lifespan)

# CORS
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[frontend_url, "http://localhost:3000", "https://build-on-code-1.preview.emergentagent.com"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============ PASSWORD UTILS ============
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

# ============ JWT UTILS ============
def get_jwt_secret() -> str:
    return os.environ.get("JWT_SECRET", "default-secret-change-me")

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(days=1),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=30),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ ADMIN SEEDING ============
async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@pokemon.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin123!")
    
    # Seed admin
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc)
        })
        print(f"Admin user created: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
    
    # Seed child account for Léo
    child_email = "maurat.leo@gmail.com"
    child_password = "Facile33"
    existing_child = await db.users.find_one({"email": child_email})
    if existing_child is None:
        # Check if old email exists and migrate
        old_child = await db.users.find_one({"email": "leo@pokemon.com"})
        if old_child:
            await db.users.update_one(
                {"email": "leo@pokemon.com"},
                {"$set": {"email": child_email, "password_hash": hash_password(child_password)}}
            )
            print(f"Child user migrated to: {child_email}")
        else:
            hashed = hash_password(child_password)
            await db.users.insert_one({
                "email": child_email,
                "password_hash": hashed,
                "name": "Léo",
                "role": "child",
                "stars": 0,
                "badges": [],
                "games_played": 0,
                "high_scores": {"memory": 0, "quiz": 0, "catch": 0},
                "created_at": datetime.now(timezone.utc)
            })
            print(f"Child user created: {child_email}")
    elif not verify_password(child_password, existing_child["password_hash"]):
        await db.users.update_one(
            {"email": child_email},
            {"$set": {"password_hash": hash_password(child_password)}}
        )
    
    # Write credentials
    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write(f"""# Test Credentials

## Admin User (Parent)
- Email: {admin_email}
- Password: {admin_password}
- Role: admin
- Permissions: Full access (add, edit, delete cards)

## Child User (Léo)
- Email: maurat.leo@gmail.com
- Password: Facile33
- Role: child
- Permissions: View cards, play games (cannot delete or edit prices)

## Auth Endpoints
- POST /api/auth/register
- POST /api/auth/login
- POST /api/auth/logout
- GET /api/auth/me
- POST /api/auth/refresh
""")

# ============ PYDANTIC MODELS ============
class UserRegister(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=2)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class CardCreate(BaseModel):
    pokemon_name: str
    card_name: str
    set_name: str
    card_number: Optional[str] = None
    image_url: str
    price: float = 0.0
    condition: str = "good"
    quantity: int = 1
    tcg_id: Optional[str] = None
    rarity: Optional[str] = None
    types: Optional[List[str]] = None
    binder_id: Optional[str] = None

class CardUpdate(BaseModel):
    price: Optional[float] = None
    condition: Optional[str] = None
    quantity: Optional[int] = None
    binder_id: Optional[str] = None

class BinderCreate(BaseModel):
    name: str = Field(min_length=1, max_length=50)
    description: Optional[str] = None
    color: str = "from-purple-500 to-pink-500"

class AIAnalyzeRequest(BaseModel):
    image_base64: str

# ============ AUTH ENDPOINTS ============
@app.post("/api/auth/register")
async def register(data: UserRegister, response: Response):
    email = data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed = hash_password(data.password)
    result = await db.users.insert_one({
        "email": email,
        "password_hash": hashed,
        "name": data.name,
        "role": "user",
        "created_at": datetime.now(timezone.utc)
    })
    
    user_id = str(result.inserted_id)
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=2592000, path="/")
    
    return {"id": user_id, "email": email, "name": data.name, "role": "user"}

@app.post("/api/auth/login")
async def login(data: UserLogin, request: Request, response: Response):
    email = data.email.lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    
    # Check brute force
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        lockout_time = attempt.get("last_attempt", datetime.now(timezone.utc))
        if datetime.now(timezone.utc) - lockout_time < timedelta(minutes=15):
            raise HTTPException(status_code=429, detail="Too many attempts. Try again in 15 minutes.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})
    
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(data.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc)}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Clear attempts on success
    await db.login_attempts.delete_one({"identifier": identifier})
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=2592000, path="/")
    
    return {"id": user_id, "email": email, "name": user.get("name", ""), "role": user.get("role", "user")}

@app.post("/api/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@app.get("/api/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@app.post("/api/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        access_token = create_access_token(str(user["_id"]), user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
        return {"message": "Token refreshed"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# Store reset codes in memory (simple approach)
reset_codes = {}

@app.post("/api/auth/forgot-password")
async def forgot_password(request: Request):
    data = await request.json()
    email = data.get("email", "").strip().lower()
    
    user = await db.users.find_one({"email": email})
    if not user:
        # Don't reveal if email exists - just say "sent"
        return {"message": "Si ce compte existe, un code a été envoyé."}
    
    # Generate 6-digit code
    code = f"{secrets.randbelow(900000) + 100000}"
    reset_codes[email] = {"code": code, "expires": datetime.now(timezone.utc).timestamp() + 600}  # 10 min
    
    print(f"[RESET] Code for {email}: {code}")
    
    return {"message": "Si ce compte existe, un code a été envoyé.", "hint": code}

@app.post("/api/auth/reset-password")
async def reset_password(request: Request):
    data = await request.json()
    email = data.get("email", "").strip().lower()
    code = data.get("code", "").strip()
    new_password = data.get("new_password", "")
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Le mot de passe doit faire au moins 6 caractères")
    
    stored = reset_codes.get(email)
    if not stored:
        raise HTTPException(status_code=400, detail="Aucune demande de réinitialisation trouvée")
    
    if datetime.now(timezone.utc).timestamp() > stored["expires"]:
        del reset_codes[email]
        raise HTTPException(status_code=400, detail="Le code a expiré")
    
    if stored["code"] != code:
        raise HTTPException(status_code=400, detail="Code incorrect")
    
    await db.users.update_one(
        {"email": email},
        {"$set": {"password_hash": hash_password(new_password)}}
    )
    
    del reset_codes[email]
    return {"message": "Mot de passe mis à jour avec succès !"}

# ============ FRENCH TO ENGLISH POKEMON NAMES ============
POKEMON_FR_TO_EN = {
    # Gen 1
    "bulbizarre": "bulbasaur", "herbizarre": "ivysaur", "florizarre": "venusaur",
    "salamèche": "charmander", "reptincel": "charmeleon", "dracaufeu": "charizard",
    "carapuce": "squirtle", "carabaffe": "wartortle", "tortank": "blastoise",
    "chenipan": "caterpie", "chrysacier": "metapod", "papilusion": "butterfree",
    "aspicot": "weedle", "coconfort": "kakuna", "dardargnan": "beedrill",
    "roucool": "pidgey", "roucoups": "pidgeotto", "roucarnage": "pidgeot",
    "rattata": "rattata", "rattatac": "raticate",
    "piafabec": "spearow", "rapasdepic": "fearow",
    "abo": "ekans", "arbok": "arbok",
    "pikachu": "pikachu", "raichu": "raichu",
    "sabelette": "sandshrew", "sablaireau": "sandslash",
    "nidoran": "nidoran", "nidorina": "nidorina", "nidoqueen": "nidoqueen",
    "nidorino": "nidorino", "nidoking": "nidoking",
    "mélofée": "clefairy", "mélodelfe": "clefable",
    "goupix": "vulpix", "feunard": "ninetales",
    "rondoudou": "jigglypuff", "grodoudou": "wigglytuff",
    "nosferapti": "zubat", "nosferalto": "golbat", "nostenfer": "crobat",
    "mystherbe": "oddish", "ortide": "gloom", "rafflesia": "vileplume",
    "paras": "paras", "parasect": "parasect",
    "mimitoss": "venonat", "aéromite": "venomoth",
    "taupiqueur": "diglett", "triopikeur": "dugtrio",
    "miaouss": "meowth", "persian": "persian",
    "psykokwak": "psyduck", "akwakwak": "golduck",
    "férosinge": "mankey", "colossinge": "primeape",
    "caninos": "growlithe", "arcanin": "arcanine",
    "ptitard": "poliwag", "têtarte": "poliwhirl", "tartard": "poliwrath",
    "abra": "abra", "kadabra": "kadabra", "alakazam": "alakazam",
    "machoc": "machop", "machopeur": "machoke", "mackogneur": "machamp",
    "chétiflor": "bellsprout", "boustiflor": "weepinbell", "empiflor": "victreebel",
    "tentacool": "tentacool", "tentacruel": "tentacruel",
    "racaillou": "geodude", "gravalanch": "graveler", "grolem": "golem",
    "ponyta": "ponyta", "galopa": "rapidash",
    "ramoloss": "slowpoke", "flagadoss": "slowbro",
    "magnéti": "magnemite", "magnéton": "magneton", "magnézone": "magnezone",
    "canarticho": "farfetch'd", "doduo": "doduo", "dodrio": "dodrio",
    "otaria": "seel", "lamantine": "dewgong",
    "tadmorv": "grimer", "grotadmorv": "muk",
    "kokiyas": "shellder", "crustabri": "cloyster",
    "fantominus": "gastly", "spectrum": "haunter", "ectoplasma": "gengar",
    "onix": "onix", "steelix": "steelix",
    "soporifik": "drowzee", "hypnomade": "hypno",
    "krabby": "krabby", "krabboss": "kingler",
    "voltorbe": "voltorb", "électrode": "electrode",
    "noeunoeuf": "exeggcute", "noadkoko": "exeggutor",
    "osselait": "cubone", "ossatueur": "marowak",
    "tygnon": "hitmonchan", "kicklee": "hitmonlee",
    "excelangue": "lickitung",
    "smogo": "koffing", "smogogo": "weezing",
    "rhinocorne": "rhyhorn", "rhinoféros": "rhydon", "rhinastoc": "rhyperior",
    "leveinard": "chansey", "leuphorie": "blissey",
    "saquedeneu": "tangela",
    "kangourex": "kangaskhan",
    "hypotrempe": "horsea", "hypocéan": "seadra", "hyporoi": "kingdra",
    "poissirène": "goldeen", "poissoroy": "seaking",
    "stari": "staryu", "staross": "starmie",
    "m. mime": "mr. mime", "mr. mime": "mr. mime",
    "insécateur": "scyther", "cizayox": "scizor",
    "lippoutou": "jynx",
    "élektek": "electabuzz", "élekable": "electivire",
    "magmar": "magmar", "maganon": "magmortar",
    "scarabrute": "pinsir",
    "tauros": "tauros",
    "magicarpe": "magikarp", "léviator": "gyarados",
    "lokhlass": "lapras",
    "métamorph": "ditto",
    "évoli": "eevee", "aquali": "vaporeon", "voltali": "jolteon", "pyroli": "flareon",
    "mentali": "espeon", "noctali": "umbreon", "givrali": "glaceon", "phyllali": "leafeon",
    "nymphali": "sylveon",
    "porygon": "porygon", "porygon2": "porygon2", "porygon-z": "porygon-z",
    "amonita": "omanyte", "amonistar": "omastar",
    "kabuto": "kabuto", "kabutops": "kabutops",
    "ptéra": "aerodactyl",
    "ronflex": "snorlax",
    "artikodin": "articuno", "électhor": "zapdos", "sulfura": "moltres",
    "minidraco": "dratini", "draco": "dragonair", "dracolosse": "dragonite",
    "mewtwo": "mewtwo", "mew": "mew",
    # Gen 2-4
    "lugia": "lugia", "ho-oh": "ho-oh", "celebi": "celebi",
    "lucario": "lucario", "riolu": "riolu",
    "arceus": "arceus", "dialga": "dialga", "palkia": "palkia", "giratina": "giratina",
    "darkrai": "darkrai", "cresselia": "cresselia",
    "tyranocif": "tyranitar", "embrylex": "larvitar", "ymphect": "pupitar",
    "métalosse": "metagross", "métang": "metang", "terhal": "beldum",
    "gardévoir": "gardevoir", "kirlia": "kirlia", "tarsal": "ralts", "gallame": "gallade",
    "clamiral": "samurott", "moustillon": "oshawott", "mateloutre": "dewott",
    # Gen 5+
    "reshiram": "reshiram", "zekrom": "zekrom", "kyurem": "kyurem",
    "xerneas": "xerneas", "yveltal": "yveltal", "zygarde": "zygarde",
    "diancie": "diancie",
    "solgaleo": "solgaleo", "lunala": "lunala", "necrozma": "necrozma",
    "zacian": "zacian", "zamazenta": "zamazenta", "éternatus": "eternatus",
    "desséliande": "trevenant", "brocélôme": "phantump",
    "rayquaza": "rayquaza", "kyogre": "kyogre", "groudon": "groudon",
    "deoxys": "deoxys", "jirachi": "jirachi",
    "feunec": "fennekin", "roussil": "braixen", "goupelin": "delphox",
    "grenousse": "froakie", "croâporal": "frogadier", "amphinobi": "greninja",
    "marisson": "chespin", "boguérisse": "quilladin", "blindépique": "chesnaught",
    "flambusard": "talonflame",
    "prismillon": "vivillon",
    "brutalibré": "hawlucha",
    "couafarel": "furfrou",
    "dedenne": "dedenne",
    "pandespiègle": "pancham", "pandarbare": "pangoro",
    # Scarlet & Violet
    "miraidon": "miraidon", "koraidon": "koraidon",
    "palafin": "palafin", "dofin": "finizen",
    "fort-ivoire": "great tusk", "fer-épine": "iron thorns",
    # Common search terms
    "sarmurai": "samurott",
}

# Prefixes and suffixes to strip for better matching
FR_PREFIXES = ["méga-", "mega-", "méga ", "mega "]
CARD_SUFFIXES = [" ex", " gx", " vmax", " vstar", " v", " tag team", " break", " lv.x", " prime"]

def translate_pokemon_name(name: str) -> str:
    """Translate French Pokemon name to English, handling prefixes/suffixes"""
    original = name.strip()
    name_lower = original.lower().strip()
    
    # Extract prefix (Méga) and suffix (EX, GX, V, VMAX...)
    prefix = ""
    suffix = ""
    base_name = name_lower
    
    for p in FR_PREFIXES:
        if base_name.startswith(p):
            prefix = "M " if "méga" in p or "mega" in p else ""
            base_name = base_name[len(p):]
            break
    
    for s in CARD_SUFFIXES:
        if base_name.endswith(s):
            suffix = s.upper().strip()
            base_name = base_name[:-len(s)].strip()
            break
    
    # Remove accents for matching
    import unicodedata
    base_clean = ''.join(c for c in unicodedata.normalize('NFD', base_name) if unicodedata.category(c) != 'Mn')
    
    # Direct match
    translated = None
    if base_name in POKEMON_FR_TO_EN:
        translated = POKEMON_FR_TO_EN[base_name]
    elif base_clean in POKEMON_FR_TO_EN:
        translated = POKEMON_FR_TO_EN[base_clean]
    else:
        # Partial match
        for fr, en in POKEMON_FR_TO_EN.items():
            fr_clean = ''.join(c for c in unicodedata.normalize('NFD', fr) if unicodedata.category(c) != 'Mn')
            if fr in base_name or base_name in fr or fr_clean in base_clean or base_clean in fr_clean:
                translated = en
                break
    
    if translated:
        result = f"{prefix}{translated} {suffix}".strip()
        return result
    
    # If no match found, return original but reconstructed
    return original

# ============ POKEMON TCG API ============
@app.get("/api/pokemon/search")
async def search_pokemon(q: str, page: int = 1, pageSize: int = 20):
    # Try French to English translation
    search_term = translate_pokemon_name(q)
    
    async with httpx.AsyncClient() as client:
        try:
            # First try: exact name with wildcard
            response = await client.get(
                f"{POKEMON_TCG_API}/cards",
                params={"q": f"name:{search_term}*", "page": page, "pageSize": pageSize},
                timeout=15.0
            )
            data = response.json()
            
            # If no results and original != translated, also try original
            if not data.get("data") and search_term.lower() != q.lower().strip():
                response = await client.get(
                    f"{POKEMON_TCG_API}/cards",
                    params={"q": f"name:{q}*", "page": page, "pageSize": pageSize},
                    timeout=15.0
                )
                data = response.json()
            
            # If still no results, try without suffix (EX, GX, etc.)
            if not data.get("data"):
                import re
                base = re.sub(r'\s*(ex|gx|vmax|vstar|v|tag team|break)\s*$', '', search_term, flags=re.IGNORECASE).strip()
                # Also strip M / Mega prefix
                base = re.sub(r'^(M|Mega)\s+', '', base, flags=re.IGNORECASE).strip()
                if base != search_term:
                    response = await client.get(
                        f"{POKEMON_TCG_API}/cards",
                        params={"q": f"name:{base}*", "page": page, "pageSize": pageSize},
                        timeout=15.0
                    )
                    data = response.json()
            
            cards = []
            for card in data.get("data", []):
                cards.append({
                    "id": card.get("id"),
                    "name": card.get("name"),
                    "set": card.get("set", {}).get("name", "Unknown"),
                    "set_id": card.get("set", {}).get("id", ""),
                    "set_total": card.get("set", {}).get("printedTotal", 0),
                    "number": card.get("number"),
                    "rarity": card.get("rarity"),
                    "types": card.get("types", []),
                    "image": card.get("images", {}).get("small"),
                    "image_large": card.get("images", {}).get("large"),
                    "prices": card.get("tcgplayer", {}).get("prices", {}),
                })
            return {"cards": cards, "totalCount": data.get("totalCount", 0)}
        except Exception as e:
            print(f"Pokemon TCG API error: {e}")
            return {"cards": [], "totalCount": 0}

@app.get("/api/pokemon/card/{card_id}")
async def get_pokemon_card(card_id: str):
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(f"{POKEMON_TCG_API}/cards/{card_id}", timeout=10.0)
            data = response.json().get("data", {})
            prices = data.get("tcgplayer", {}).get("prices", {})
            market_price = 0
            for variant, price_data in prices.items():
                if "market" in price_data:
                    market_price = price_data["market"]
                    break
            return {
                "id": data.get("id"),
                "name": data.get("name"),
                "set": data.get("set", {}).get("name"),
                "number": data.get("number"),
                "rarity": data.get("rarity"),
                "image": data.get("images", {}).get("large"),
                "prices": prices,
                "market_price": market_price
            }
        except Exception as e:
            raise HTTPException(status_code=404, detail="Card not found")

# ============ AI CARD RECOGNITION ============
@app.post("/api/ai/analyze-card")
async def analyze_card(data: AIAnalyzeRequest, request: Request):
    user = await get_current_user(request)
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="AI service not configured")
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"card-analysis-{user['_id']}-{datetime.now().timestamp()}",
            system_message="""You are a world-class Pokemon TCG card identification expert. You can identify ANY Pokemon card from a photo with extreme precision.

When you see a Pokemon card image, you MUST extract:
1. **English name** of the Pokemon (even if the card is in French, Japanese, etc.)
2. **Set name** in English (e.g. "Paldean Fates", "Obsidian Flames", "Crown Zenith")
3. **Card number** (e.g. "057/091", "25/198")
4. **Rarity** (Common, Uncommon, Rare, Holo Rare, Ultra Rare, Secret Rare, etc.)
5. **Card condition estimate** (mint, excellent, good, poor)
6. **French name** of the Pokemon if the card is in French
7. **Card suffix** if applicable (ex, GX, V, VMAX, VSTAR, EX)

IMPORTANT IDENTIFICATION TIPS:
- Look at the set symbol (bottom right or left of the card art)
- Read the card number at the bottom (format: XXX/YYY)
- Read the Pokemon name at the top of the card
- The set can often be identified by the set symbol icon
- For French cards: "Ectoplasma" = Gengar, "Dracaufeu" = Charizard, etc.
- For EX/GX/V cards, include the suffix in the name

Respond ONLY in this exact JSON format (no extra text):
{"pokemon_name_en": "Gengar", "pokemon_name_fr": "Ectoplasma", "set_name": "Paldean Fates", "card_number": "057/091", "rarity": "Holo Rare", "condition": "good", "suffix": "ex"}

If you cannot identify something with certainty, provide your best guess. NEVER use null - always try."""
        ).with_model("openai", "gpt-4o")
        
        image_content = ImageContent(image_base64=data.image_base64)
        user_message = UserMessage(
            text="Identify this Pokemon card precisely. Extract the English name, French name if applicable, set name, card number, rarity and condition.",
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        
        # Parse JSON from response
        import json as json_module
        import re
        json_match = re.search(r'\{[^}]+\}', response)
        if not json_match:
            return {"pokemon_name_en": None, "pokemon_name_fr": None, "set_name": None, "card_number": None, "rarity": None, "condition": "good", "tcg_matches": []}
        
        ai_result = json_module.loads(json_match.group())
        
        # Now search Pokemon TCG API for exact match
        tcg_matches = []
        async with httpx.AsyncClient() as client:
            en_name = ai_result.get("pokemon_name_en", "")
            suffix = ai_result.get("suffix", "")
            card_number = ai_result.get("card_number", "")
            set_name = ai_result.get("set_name", "")
            
            # Build precise search query
            search_name = en_name
            if suffix:
                search_name = f"{en_name} {suffix}".strip()
            
            queries_to_try = []
            
            # Most precise: name + number
            if card_number and "/" in card_number:
                num = card_number.split("/")[0].lstrip("0")
                queries_to_try.append(f'name:"{search_name}" number:{num}')
            
            # Name + set
            if set_name:
                queries_to_try.append(f'name:"{search_name}" set.name:"{set_name}"')
            
            # Just the name with suffix
            queries_to_try.append(f'name:"{search_name}"')
            
            # Fallback: base name wildcard
            queries_to_try.append(f'name:{en_name}*')
            
            for q in queries_to_try:
                try:
                    resp = await client.get(
                        f"{POKEMON_TCG_API}/cards",
                        params={"q": q, "pageSize": 10},
                        timeout=10.0
                    )
                    tcg_data = resp.json()
                    if tcg_data.get("data"):
                        for card in tcg_data["data"]:
                            tcg_matches.append({
                                "id": card.get("id"),
                                "name": card.get("name"),
                                "set": card.get("set", {}).get("name", ""),
                                "set_id": card.get("set", {}).get("id", ""),
                                "number": card.get("number"),
                                "rarity": card.get("rarity"),
                                "types": card.get("types", []),
                                "image": card.get("images", {}).get("small"),
                                "image_large": card.get("images", {}).get("large"),
                                "prices": card.get("tcgplayer", {}).get("prices", {}),
                            })
                        break  # Found results, stop trying
                except:
                    continue
        
        ai_result["tcg_matches"] = tcg_matches
        return ai_result
        
    except Exception as e:
        print(f"AI analysis error: {e}")
        raise HTTPException(status_code=500, detail=f"AI analysis failed: {str(e)}")

# ============ CARD COLLECTION CRUD ============
@app.get("/api/cards")
async def get_cards(request: Request):
    user = await get_current_user(request)
    
    # Children see cards from admin/parent AND their own
    if user.get("role") == "child":
        admin = await db.users.find_one({"role": "admin"})
        if admin:
            cursor = db.cards.find({"user_id": {"$in": [str(admin["_id"]), user["_id"]]}}).sort("created_at", -1)
        else:
            cursor = db.cards.find({"user_id": user["_id"]}).sort("created_at", -1)
    else:
        cursor = db.cards.find({"user_id": user["_id"]}).sort("created_at", -1)
    
    cards = []
    async for card in cursor:
        card["_id"] = str(card["_id"])
        cards.append(card)
    return {"cards": cards}

@app.post("/api/cards")
async def create_card(data: CardCreate, request: Request):
    user = await get_current_user(request)
    card_doc = {
        "user_id": user["_id"],
        "pokemon_name": data.pokemon_name,
        "card_name": data.card_name,
        "set_name": data.set_name,
        "card_number": data.card_number,
        "image_url": data.image_url,
        "price": data.price,
        "condition": data.condition,
        "quantity": data.quantity,
        "tcg_id": data.tcg_id,
        "rarity": data.rarity,
        "types": data.types or [],
        "binder_id": data.binder_id,
        "created_at": datetime.now(timezone.utc),
        "price_history": [{"price": data.price, "date": datetime.now(timezone.utc).isoformat()}]
    }
    result = await db.cards.insert_one(card_doc)
    card_doc["_id"] = str(result.inserted_id)
    return card_doc

@app.put("/api/cards/{card_id}")
async def update_card(card_id: str, request: Request):
    user = await get_current_user(request)
    
    # Children cannot update cards
    if user.get("role") == "child":
        raise HTTPException(status_code=403, detail="Les enfants ne peuvent pas modifier les cartes")
    
    card = await db.cards.find_one({"_id": ObjectId(card_id), "user_id": user["_id"]})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    
    raw_data = await request.json()
    update_data = {}
    
    if "price" in raw_data and raw_data["price"] is not None:
        update_data["price"] = float(raw_data["price"])
        update_data["$push"] = {"price_history": {"price": float(raw_data["price"]), "date": datetime.now(timezone.utc).isoformat()}}
    if "condition" in raw_data and raw_data["condition"] is not None:
        update_data["condition"] = raw_data["condition"]
    if "quantity" in raw_data and raw_data["quantity"] is not None:
        update_data["quantity"] = int(raw_data["quantity"])
    if "binder_id" in raw_data:
        update_data["binder_id"] = raw_data["binder_id"]  # Can be None to remove from binder
    
    if "$push" in update_data:
        push_data = update_data.pop("$push")
        await db.cards.update_one(
            {"_id": ObjectId(card_id)},
            {"$set": update_data, "$push": push_data}
        )
    else:
        await db.cards.update_one({"_id": ObjectId(card_id)}, {"$set": update_data})
    
    updated = await db.cards.find_one({"_id": ObjectId(card_id)})
    updated["_id"] = str(updated["_id"])
    return updated

@app.delete("/api/cards/{card_id}")
async def delete_card(card_id: str, request: Request):
    user = await get_current_user(request)
    
    # Children cannot delete cards
    if user.get("role") == "child":
        raise HTTPException(status_code=403, detail="Les enfants ne peuvent pas supprimer les cartes")
    
    result = await db.cards.delete_one({"_id": ObjectId(card_id), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Card not found")
    return {"message": "Card deleted"}

# ============ STATISTICS ============
@app.get("/api/stats")
async def get_stats(request: Request):
    user = await get_current_user(request)
    
    # Children see stats from admin/parent
    if user.get("role") == "child":
        admin = await db.users.find_one({"role": "admin"})
        target_user_id = str(admin["_id"]) if admin else user["_id"]
    else:
        target_user_id = user["_id"]
    
    pipeline = [
        {"$match": {"user_id": target_user_id}},
        {"$group": {
            "_id": None,
            "total_cards": {"$sum": "$quantity"},
            "total_value": {"$sum": {"$multiply": ["$price", "$quantity"]}},
            "unique_pokemon": {"$addToSet": "$pokemon_name"},
            "mint_count": {"$sum": {"$cond": [{"$eq": ["$condition", "mint"]}, "$quantity", 0]}},
            "max_price": {"$max": "$price"}
        }}
    ]
    
    result = await db.cards.aggregate(pipeline).to_list(1)
    
    if not result:
        return {
            "total_cards": 0,
            "total_value": 0,
            "unique_pokemon": 0,
            "avg_value": 0,
            "mint_count": 0,
            "top_card": None
        }
    
    stats = result[0]
    total_cards = stats.get("total_cards", 0)
    total_value = stats.get("total_value", 0)
    
    # Get top card
    top_card = await db.cards.find_one(
        {"user_id": target_user_id},
        sort=[("price", -1)]
    )
    
    return {
        "total_cards": total_cards,
        "total_value": round(total_value, 2),
        "unique_pokemon": len(stats.get("unique_pokemon", [])),
        "avg_value": round(total_value / total_cards, 2) if total_cards > 0 else 0,
        "mint_count": stats.get("mint_count", 0),
        "top_card": {
            "name": top_card.get("card_name") if top_card else None,
            "price": top_card.get("price") if top_card else 0
        }
    }

@app.get("/api/stats/top-cards")
async def get_top_cards(request: Request, limit: int = 5):
    user = await get_current_user(request)
    
    # Children see stats from admin/parent
    if user.get("role") == "child":
        admin = await db.users.find_one({"role": "admin"})
        target_user_id = str(admin["_id"]) if admin else user["_id"]
    else:
        target_user_id = user["_id"]
    
    cursor = db.cards.find({"user_id": target_user_id}).sort("price", -1).limit(limit)
    cards = []
    async for card in cursor:
        card["_id"] = str(card["_id"])
        cards.append(card)
    return {"cards": cards}

# ============ SHARE COLLECTION ============
@app.post("/api/share")
async def create_share_link(request: Request):
    user = await get_current_user(request)
    share_token = secrets.token_urlsafe(16)
    
    await db.shares.update_one(
        {"user_id": user["_id"]},
        {"$set": {
            "token": share_token,
            "user_name": user.get("name", "Collector"),
            "created_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    
    return {"share_token": share_token}

@app.get("/api/share/{token}")
async def get_shared_collection(token: str):
    share = await db.shares.find_one({"token": token})
    if not share:
        raise HTTPException(status_code=404, detail="Share not found")
    
    cursor = db.cards.find({"user_id": share["user_id"]}).sort("price", -1)
    cards = []
    async for card in cursor:
        card["_id"] = str(card["_id"])
        cards.append(card)
    
    # Stats
    total_value = sum(c["price"] * c["quantity"] for c in cards)
    total_cards = sum(c["quantity"] for c in cards)
    
    return {
        "collector_name": share.get("user_name", "Collector"),
        "cards": cards,
        "stats": {
            "total_cards": total_cards,
            "total_value": round(total_value, 2),
            "unique_pokemon": len(set(c["pokemon_name"] for c in cards))
        }
    }

# ============ EXPORT ============
@app.get("/api/export/pdf")
async def export_pdf(request: Request):
    user = await get_current_user(request)
    
    cursor = db.cards.find({"user_id": user["_id"]}).sort("pokemon_name", 1)
    cards = await cursor.to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph(f"Collection Pokemon - {user.get('name', 'Collector')}", styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [["Pokemon", "Set", "Condition", "Qty", "Price", "Total"]]
    total_value = 0
    for card in cards:
        total = card["price"] * card["quantity"]
        total_value += total
        data.append([
            card["pokemon_name"],
            card["set_name"][:20],
            card["condition"],
            str(card["quantity"]),
            f"{card['price']:.2f}€",
            f"{total:.2f}€"
        ])
    
    data.append(["", "", "", "", "TOTAL:", f"{total_value:.2f}€"])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF007F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#121218')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2D2D3B')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=collection_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

@app.get("/api/export/excel")
async def export_excel(request: Request):
    user = await get_current_user(request)
    
    cursor = db.cards.find({"user_id": user["_id"]}).sort("pokemon_name", 1)
    cards = await cursor.to_list(1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection"
    
    # Headers
    headers = ["Pokemon", "Card Name", "Set", "Number", "Rarity", "Condition", "Quantity", "Price (€)", "Total (€)"]
    ws.append(headers)
    
    # Style header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="FF007F", end_color="FF007F", fill_type="solid")
    
    # Data
    total_value = 0
    for card in cards:
        total = card["price"] * card["quantity"]
        total_value += total
        ws.append([
            card["pokemon_name"],
            card["card_name"],
            card["set_name"],
            card.get("card_number", ""),
            card.get("rarity", ""),
            card["condition"],
            card["quantity"],
            card["price"],
            total
        ])
    
    # Total row
    ws.append(["", "", "", "", "", "", "", "TOTAL:", total_value])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=collection_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ============ BINDERS / CLASSEURS ============
@app.get("/api/binders")
async def get_binders(request: Request):
    user = await get_current_user(request)
    target_id = user["_id"]
    if user.get("role") == "child":
        admin = await db.users.find_one({"role": "admin"})
        if admin:
            target_id = str(admin["_id"])
    
    cursor = db.binders.find({"user_id": target_id}).sort("created_at", -1)
    binders = []
    async for b in cursor:
        b["_id"] = str(b["_id"])
        binders.append(b)
    return {"binders": binders}

@app.post("/api/binders")
async def create_binder(data: BinderCreate, request: Request):
    user = await get_current_user(request)
    if user.get("role") == "child":
        raise HTTPException(status_code=403, detail="Les enfants ne peuvent pas créer de classeurs")
    
    binder_doc = {
        "user_id": user["_id"],
        "name": data.name,
        "description": data.description,
        "color": data.color,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.binders.insert_one(binder_doc)
    binder_doc["_id"] = str(result.inserted_id)
    return binder_doc

@app.put("/api/binders/{binder_id}")
async def update_binder(binder_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") == "child":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    data = await request.json()
    update_data = {}
    if "name" in data:
        update_data["name"] = data["name"]
    if "description" in data:
        update_data["description"] = data["description"]
    if "color" in data:
        update_data["color"] = data["color"]
    
    if update_data:
        await db.binders.update_one(
            {"_id": ObjectId(binder_id), "user_id": user["_id"]},
            {"$set": update_data}
        )
    
    updated = await db.binders.find_one({"_id": ObjectId(binder_id)})
    if not updated:
        raise HTTPException(status_code=404, detail="Classeur non trouvé")
    updated["_id"] = str(updated["_id"])
    return updated

@app.delete("/api/binders/{binder_id}")
async def delete_binder(binder_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") == "child":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    result = await db.binders.delete_one({"_id": ObjectId(binder_id), "user_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Classeur non trouvé")
    
    # Remove binder_id from cards in this binder
    await db.cards.update_many(
        {"binder_id": binder_id},
        {"$set": {"binder_id": None}}
    )
    return {"message": "Classeur supprimé"}

# ============ CARDS BY SET (for set grouping view) ============
@app.get("/api/cards/by-set")
async def get_cards_by_set(request: Request):
    user = await get_current_user(request)
    target_id = user["_id"]
    if user.get("role") == "child":
        admin = await db.users.find_one({"role": "admin"})
        if admin:
            target_id = str(admin["_id"])
    
    pipeline = [
        {"$match": {"user_id": target_id}},
        {"$group": {
            "_id": "$set_name",
            "count": {"$sum": "$quantity"},
            "total_value": {"$sum": {"$multiply": ["$price", "$quantity"]}},
            "cards": {"$push": {
                "id": {"$toString": "$_id"},
                "pokemon_name": "$pokemon_name",
                "card_name": "$card_name",
                "image_url": "$image_url",
                "price": "$price",
                "rarity": "$rarity",
                "types": "$types"
            }}
        }},
        {"$sort": {"count": -1}}
    ]
    
    sets = await db.cards.aggregate(pipeline).to_list(100)
    return {"sets": [{"set_name": s["_id"], "count": s["count"], "total_value": round(s["total_value"], 2), "cards": s["cards"]} for s in sets]}

# ============ VITRINE (Public Showcase) ============
@app.post("/api/vitrine/create")
async def create_vitrine(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Seul l'admin peut créer une vitrine")
    
    data = await request.json()
    vitrine_token = secrets.token_urlsafe(16)
    
    await db.vitrines.update_one(
        {"user_id": user["_id"]},
        {"$set": {
            "token": vitrine_token,
            "title": data.get("title", "Ma Collection Pokémon"),
            "description": data.get("description", ""),
            "user_name": user.get("name", "Collectionneur"),
            "created_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    return {"token": vitrine_token}

@app.get("/api/vitrine/{token}")
async def get_vitrine(token: str):
    vitrine = await db.vitrines.find_one({"token": token})
    if not vitrine:
        raise HTTPException(status_code=404, detail="Vitrine non trouvée")
    
    cursor = db.cards.find({"user_id": vitrine["user_id"]}).sort("price", -1)
    cards = []
    async for card in cursor:
        cards.append({
            "pokemon_name": card.get("pokemon_name"),
            "card_name": card.get("card_name"),
            "set_name": card.get("set_name"),
            "image_url": card.get("image_url"),
            "price": card.get("price", 0),
            "condition": card.get("condition"),
            "rarity": card.get("rarity"),
            "types": card.get("types", []),
            "quantity": card.get("quantity", 1)
        })
    
    total_value = sum(c["price"] * c["quantity"] for c in cards)
    total_cards = sum(c["quantity"] for c in cards)
    
    # Group by type
    type_counts = {}
    for c in cards:
        for t in c.get("types", []):
            type_counts[t] = type_counts.get(t, 0) + c["quantity"]
    
    # Group by set
    set_counts = {}
    for c in cards:
        sn = c.get("set_name", "Inconnu")
        if sn not in set_counts:
            set_counts[sn] = 0
        set_counts[sn] += c["quantity"]
    
    return {
        "title": vitrine.get("title", "Ma Collection"),
        "description": vitrine.get("description", ""),
        "collector_name": vitrine.get("user_name", "Collectionneur"),
        "cards": cards,
        "stats": {
            "total_cards": total_cards,
            "total_value": round(total_value, 2),
            "unique_pokemon": len(set(c["pokemon_name"] for c in cards)),
            "type_distribution": type_counts,
            "set_distribution": dict(sorted(set_counts.items(), key=lambda x: x[1], reverse=True))
        }
    }

# ============ HEALTH CHECK ============
@app.get("/api/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# ============ GAMES & REWARDS ============
import random

# Pokemon data for games
POKEMON_FOR_GAMES = [
    {"name": "Pikachu", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/25.png"},
    {"name": "Dracaufeu", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/6.png"},
    {"name": "Tortank", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/9.png"},
    {"name": "Florizarre", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/3.png"},
    {"name": "Mewtwo", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/150.png"},
    {"name": "Évoli", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/133.png"},
    {"name": "Ronflex", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/143.png"},
    {"name": "Dracolosse", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/149.png"},
    {"name": "Mew", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/151.png"},
    {"name": "Lucario", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/448.png"},
    {"name": "Salamèche", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/4.png"},
    {"name": "Carapuce", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/7.png"},
    {"name": "Bulbizarre", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/1.png"},
    {"name": "Raichu", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/26.png"},
    {"name": "Artikodin", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/144.png"},
    {"name": "Électhor", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/145.png"},
    {"name": "Sulfura", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/146.png"},
    {"name": "Léviator", "image": "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/130.png"},
]

BADGES = [
    {"id": "first_game", "name": "Premier Pas", "emoji": "🎮", "desc": "Joue ton premier jeu"},
    {"id": "memory_master", "name": "Maître de la Mémoire", "emoji": "🧠", "desc": "Gagne 5 parties de Memory"},
    {"id": "quiz_champion", "name": "Champion du Quiz", "emoji": "🏆", "desc": "Réponds juste 10 fois au Quiz"},
    {"id": "catcher", "name": "Attrapeur", "emoji": "🎯", "desc": "Attrape 20 Pokémon"},
    {"id": "star_collector", "name": "Collectionneur d'Étoiles", "emoji": "⭐", "desc": "Gagne 50 étoiles"},
    {"id": "super_player", "name": "Super Joueur", "emoji": "🌟", "desc": "Joue 10 parties"},
]

@app.get("/api/games/memory")
async def get_memory_game(request: Request):
    """Get cards for memory game"""
    user = await get_current_user(request)
    
    # Select 6 random pokemon (12 cards total for pairs)
    selected = random.sample(POKEMON_FOR_GAMES, 6)
    cards = []
    for i, poke in enumerate(selected):
        cards.append({"id": i*2, "pokemon": poke["name"], "image": poke["image"], "pairId": i})
        cards.append({"id": i*2+1, "pokemon": poke["name"], "image": poke["image"], "pairId": i})
    
    random.shuffle(cards)
    return {"cards": cards}

@app.get("/api/games/quiz")
async def get_quiz_question(request: Request):
    """Get a quiz question"""
    user = await get_current_user(request)
    
    correct = random.choice(POKEMON_FOR_GAMES)
    wrong_choices = random.sample([p for p in POKEMON_FOR_GAMES if p["name"] != correct["name"]], 3)
    
    options = [correct["name"]] + [p["name"] for p in wrong_choices]
    random.shuffle(options)
    
    return {
        "image": correct["image"],
        "options": options,
        "correct": correct["name"]
    }

@app.get("/api/games/catch")
async def get_catch_game(request: Request):
    """Get pokemon for catch game"""
    user = await get_current_user(request)
    
    # Return 8 pokemon to catch
    selected = random.sample(POKEMON_FOR_GAMES, 8)
    return {"pokemon": selected, "timeLimit": 30}

@app.post("/api/games/score")
async def save_game_score(request: Request):
    """Save game score and award stars"""
    user = await get_current_user(request)
    data = await request.json()
    
    game = data.get("game")  # memory, quiz, catch
    score = data.get("score", 0)
    won = data.get("won", False)
    
    # Calculate stars earned
    stars_earned = 0
    if won:
        if game == "memory":
            stars_earned = 3
        elif game == "quiz":
            stars_earned = 1
        elif game == "catch":
            stars_earned = max(1, score // 2)
    
    # Update user stats
    update_ops = {
        "$inc": {
            "stars": stars_earned,
            "games_played": 1,
            f"game_stats.{game}_played": 1,
        }
    }
    
    if won:
        update_ops["$inc"][f"game_stats.{game}_won"] = 1
    
    # Update high score if better
    current_user = await db.users.find_one({"_id": ObjectId(user["_id"])})
    high_scores = current_user.get("high_scores", {})
    if score > high_scores.get(game, 0):
        update_ops["$set"] = {f"high_scores.{game}": score}
    
    await db.users.update_one({"_id": ObjectId(user["_id"])}, update_ops)
    
    # Check for new badges
    new_badges = []
    current_badges = current_user.get("badges", [])
    game_stats = current_user.get("game_stats", {})
    current_stars = current_user.get("stars", 0) + stars_earned
    games_played = current_user.get("games_played", 0) + 1
    
    # First game badge
    if games_played == 1 and "first_game" not in current_badges:
        new_badges.append("first_game")
    
    # Memory master (5 wins)
    if game_stats.get("memory_won", 0) + (1 if won and game == "memory" else 0) >= 5 and "memory_master" not in current_badges:
        new_badges.append("memory_master")
    
    # Quiz champion (10 correct)
    if game_stats.get("quiz_won", 0) + (1 if won and game == "quiz" else 0) >= 10 and "quiz_champion" not in current_badges:
        new_badges.append("quiz_champion")
    
    # Star collector (50 stars)
    if current_stars >= 50 and "star_collector" not in current_badges:
        new_badges.append("star_collector")
    
    # Super player (10 games)
    if games_played >= 10 and "super_player" not in current_badges:
        new_badges.append("super_player")
    
    if new_badges:
        await db.users.update_one(
            {"_id": ObjectId(user["_id"])},
            {"$push": {"badges": {"$each": new_badges}}}
        )
    
    # Get badge details for new badges
    new_badge_details = [b for b in BADGES if b["id"] in new_badges]
    
    return {
        "stars_earned": stars_earned,
        "total_stars": current_stars,
        "new_badges": new_badge_details
    }

@app.get("/api/user/profile")
async def get_user_profile(request: Request):
    """Get user profile with stars and badges"""
    user = await get_current_user(request)
    
    full_user = await db.users.find_one({"_id": ObjectId(user["_id"])})
    
    user_badges = full_user.get("badges", [])
    badge_details = [b for b in BADGES if b["id"] in user_badges]
    
    return {
        "name": full_user.get("name"),
        "role": full_user.get("role"),
        "stars": full_user.get("stars", 0),
        "badges": badge_details,
        "games_played": full_user.get("games_played", 0),
        "high_scores": full_user.get("high_scores", {}),
        "game_stats": full_user.get("game_stats", {}),
        "avatar": full_user.get("avatar")
    }

@app.post("/api/user/avatar")
async def update_avatar(request: Request):
    """Update user avatar (base64 image)"""
    user = await get_current_user(request)
    data = await request.json()
    
    avatar = data.get("avatar")  # base64 image
    if not avatar:
        raise HTTPException(status_code=400, detail="Avatar required")
    
    await db.users.update_one(
        {"_id": ObjectId(user["_id"])},
        {"$set": {"avatar": avatar}}
    )
    
    return {"message": "Avatar updated", "avatar": avatar}

@app.post("/api/user/child-avatar")
async def update_child_avatar(request: Request):
    """Admin can update child's avatar"""
    user = await get_current_user(request)
    
    # Only admin can update child avatar
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update child avatar")
    
    data = await request.json()
    child_email = data.get("child_email", "maurat.leo@gmail.com")
    avatar = data.get("avatar")
    
    if not avatar:
        raise HTTPException(status_code=400, detail="Avatar required")
    
    result = await db.users.update_one(
        {"email": child_email, "role": "child"},
        {"$set": {"avatar": avatar}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Child not found")
    
    return {"message": "Child avatar updated"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
